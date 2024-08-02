import discord
from discord.ext import commands
from discord.utils import get
from dotenv import load_dotenv
from collections import Counter
import openpyxl
import xlsxwriter
import time
import array
import os
import re
import pandas as pd

intents = discord.Intents.default()
intents.typing = False
intents.presences = False
intents.message_content = True
user_array = []
lookup_array = []
new_message = ""
numberofthumbsup = 0
bot = commands.Bot(command_prefix='!', intents=intents)
role_id = 1249201320884305983
stuff = 0
pattern = re.compile(r'http\S+|www\S+')


@bot.event
async def on_ready():
    with open("usertextfile.txt",'r') as file:
        for line in file.readlines():
            #user_array.append(line.strip())   
            print("") 
    print(f'{bot.user} has connected to Discord!')

emoji1 = bot.get_emoji('👍')
flex = bot.get_emoji('<:Flex:1199972735401607208>')
ping_message = ""
laughingatnoob = discord.utils.get(bot.emojis, name='pointandlaugh')
    

@bot.event
async def on_message(message):
    global stuff
    stuff = message
    global new_message
    global numberofthumbsup
    global ping_message
    global lookup_array
    guild = message.guild 
    guild_get = bot.get_guild(guild.id)
    #roles = guild.roles
    #needed_role = get(guild.roles, name="Guild_League_Ping")
    if bot.user.mentioned_in(message):
        await message.channel.send("I am not programmed yet to know what you said, but if its gem you are noob :).\nIf not I'm sure you are a lovely person and if you said something mean go step on a lego.")
    if message.content.startswith('!AB'):
        new_message = message.content.replace('!AB ','')
        new_message =  new_message.lower()
        #print(new_message)
        trigger_message = new_message[0:6]
        display_message = new_message[7:]
        if trigger_message == "help":
            await message.channel.send('To look something up -> !AB lookup <stuff listed in important shit>\nGo here for more help -> https://discord.com/channels/808051243183767582/1215463505017176124/1266138009175199787')
        if trigger_message == "glstar":
            messageID = message
            #print(messageID)
            ping_message = await message.channel.send('\n👍: BE READY FOR THE FIRST ROUND\n\n👎: YOU ARENT GOING\n\n⛺: MEANS TENTATIVE.. UNCERTAIN; SUBJECT TO FUTURE CHANGE. IF YOU CANT MAKE IT AT FIRST ROUND PICK THIS\n\n<:Flex:1199972735401607208>: MEANS YOU CAN MAKE IT TO THE FIRST ROUND BUT ARE THERE TO FILL\nhttps://cdn.discordapp.com/attachments/828479988948533299/1236084210519773336/caption.gif?ex=667b46a1&is=6679f521&hm=b1dfa876dfc3f642e1e4da329b31b343b9cf0352a48dfadfa3a7b8bc34d87b0c&\nGuild League' + display_message)
            await ping_message.pin()
            await ping_message.add_reaction('👍')
            await ping_message.add_reaction('👎')
            await ping_message.add_reaction('⛺')
            await ping_message.add_reaction('<:Flex:1199972735401607208>')
        if trigger_message == "reload":
            print(ping_message)
        if trigger_message == "delete":
            ping_message = await message.channel.send('volc protocol starting.. DELETING MESSAGES BEEP BOOP')
            await message.channel.purge(limit=None, check=lambda msg: msg.pinned)
            await ping_message.delete()
            channels = message.channel
            async for msg in channels.history():
                if msg.type is discord.MessageType.pins_add:
                    await msg.delete()
                if(msg.content[0:3]) == "!AB":
                    await msg.delete()
                if msg.author.name == "Angry_Bot":
                    await msg.delete()
        if trigger_message == "status":
            numberofthumbsup = len(user_array) - 1
            ping_message = await message.channel.send(str(numberofthumbsup) + "/10")
        if trigger_message == "update":
            channels = guild.channels
            workbook = xlsxwriter.Workbook(display_message+'.xlsx')
            worksheet = workbook.add_worksheet()
            display_message = new_message[7:]
            print(display_message)
            for channel in channels:
                print(channel.name)
                if(channel.name != "NetSlum"):
                    if(channel.name != "BDO Stuff"):
                        if(channel.name != "NSFW"):
                            if(channel.name != "Voice Channels"):
                                if(channel.name != "archive"):
                                    if(channel.name != "war-shit"):
                                        if(channel.name != "Welcome"):
                                            if(channel.name != "General"):
                                                if(channel.name != "class-help"):
                                                    i = 1
                                                    async for msg in channel.history(): 
                                                        if channel.name == str(display_message): 
                                                           stuff = 'A'+ str(i) 
                                                           stuff2 = 'B' + str(i)
                                                           stuff3 = 'C' + str(i)
                                                           worksheet.write(stuff, str(msg.id)) 
                                                           worksheet.write(stuff2, str(channel.name))
                                                           worksheet.write(stuff3, str(re.sub(pattern,'',msg.content)))
                                                           i = i + 1   
            workbook.close()         
            ping_message = await message.channel.send('Done :3')                             
        if trigger_message == "lookup":
            channels = guild.channels
            dm = new_message[7:26] 
            #threads = await channels.threads()
            #ping_message = await message.channel.fetch_message(msg_id)
            #await message.channel.send("stuff" + message)
            files = [f for f in os.listdir('C:\\Users\\vwalk\\Discord_bot') if f.endswith('.xlsx')]
            for file in files:
                file_path = os.path.join('C:\\Users\\vwalk\\Discord_bot', file)
                df = openpyxl.load_workbook(file_path)
                sh = df.active
                max_rows = sh.max_row
                for move in range(1,max_rows+1):
                    valuez = str(sh.cell(row=move, column=3).value)
                    valuez = valuez.lstrip()
                    if str(valuez).lower() == display_message.lower():
                        print(f"Found row {move} with value {sh.cell(row=move, column=3).value}")   
                        dm = sh.cell(row=move, column=1).value
            for channel in channels:
                #if channel.name == "bdo-resources":
                if(channel.name != "NetSlum"):
                    if(channel.name != "BDO Stuff"):
                        if(channel.name != "NSFW"):
                            if(channel.name != "Voice Channels"):
                                if(channel.name != "archive"):
                                    if(channel.name != "war-shit"):
                                        if(channel.name != "Welcome"):
                                            if(channel.name != "General"):
                                                if(channel.name != "class-help"):
                                                    async for msg in channel.history():
                                                        if str(msg.id) == str(dm):  
                                                            await message.channel.send(msg.jump_url)
                                                            return
            ping_message = await message.channel.send('Nope cant find ' + display_message + '\nIf you need help go here -> https://discord.com/channels/808051243183767582/1215463505017176124/1266138009175199787')
@bot.event 
async def on_reaction_add(reaction,user):
    global user_array
    global numberofthumbsup
    global stuff
    global flag
    if reaction.message.author == bot.user:
        if reaction.emoji == '👍':
            if user.name in user_array: 
                print("")
            else: 
                user_array.append(user.name)
                print(user_array)
                with open('usertextfile.txt', 'r') as f:
                    lines = list(filter(lambda x: x.strip() != '', f.readlines()))
                with open('usertextfile.txt', 'w') as f:
                    f.writelines(lines)
                f.close
                f = open("usertextfile.txt","w+")
                for i in user_array:
                    f.write(i+"\n")
                f.close
            if len(user_array) > 1:
                await ping_message.remove_reaction(reaction.emoji, bot.user)
            numberofthumbsup = len(user_array) - 1
            member = reaction.message.author
            role = get(member.guild.roles, id=role_id)
            await user.add_roles(role)
@bot.event
async def on_raw_reaction_remove(reaction):
    user = await bot.fetch_user(reaction.user_id)
    global user_array
    global numberofthumbsup
    global flag
    numberofthumbsup = len(user_array) - 1
    print(numberofthumbsup)
    if numberofthumbsup > 0:
        if str(user) in user_array:
            user_array.remove(str(user))
            print(len(user_array))
            if len(user_array) <= 1:
                await ping_message.add_reaction(reaction.emoji)
            print(user_array)
            numberofthumbsup = len(user_array) - 1
            f = open("usertextfile.txt","w+")
            for i in user_array:
                f.write("\n" + i)
            f.close
            with open('usertextfile.txt', 'r') as f:
                lines = list(filter(lambda x: x.strip() != '', f.readlines()))
            with open('usertextfile.txt', 'w') as f:
                f.writelines(lines)
            f.close
load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')

bot.run(TOKEN)